# https://www.bilibili.com/video/av23697305/
# 使用python操作excel，共5集
# -------------------------------------------------------------------------------------------

import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws.title

ws["A1"] = 520
ws.append([1,2,3])

import datetime
ws["A3"] = datetime.datetime.now()
wb.save("demo.xlsx")

# -------------------------------------------------------------------------------------------

import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\Administrator\Desktop\aa.xlsx")
type(wb)
wb.get_sheet_names()   #deprecated function
print(wb.sheetnames)
ws = wb.get_sheet_by_name["Sheet1"]

wb.create_sheet(index = 0,title = "FishC")
wb.remove_sheet(wb.get_sheet_by_name("FishC"))

c = ws["A2"]
print(c.row)
print(c.column)
print(c.coordinate)
print(ws["A2"].value)
print(c.value)
d = c.offset(2,0)

openpyxl.cell.cell.get_column_letter(496)
openpyxl.cell.cell.column_index_from_string("JB")

for each_movie in ws["A2":"B10"]:
    for each_cell in each_movie:
        print(each_cell.value,end = " ")
    print("\n")

for each_row in ws.rows:
    print(each_row[0].value)

for each_row in ws.iter_rows(min_row=2,min_col=1,max_row=4,max_col=2):
    print(each_row[0].value)

new = wb.copy_worksheet(ws)
type(new)
wb.save(r"C:\Users\Administrator\Desktop\bb.xlsx")

# -------------------------------------------------------------------------------------------

import openpyxl

wb = openpyxl.Workbook()

ws1 = wb.create_sheet(title = "小甲鱼")
ws1.sheet_properties.tabColor = "FF0000"

ws1.row_dimensions[2].height = 100
ws1.column_dimensions["C"].width = 50

ws1.merge_cells("A1:C3")
ws1["A1"] = "good"
ws1.unmerge_cells("A1:C3")

ws1.freeze_panes = "B8"
ws1.freeze_panes = "A1"  #或设置为none

# -------------------------------------------------------------------------------------------

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import GradientFill
from openpyxl.styles import Border,Side
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle

wb = Workbook()
ws = wb.active
b2 = ws["B2"]
b2.value = "FishC"
bold_red_font = Font(bold=True,color="FF0000")
b2.font = bold_red_font

b3 = ws["B3"]
b3.value = "Fish"
Ita_strike_font = Font(size=16,italic=True,strike=True,color="0000FF")
b3.font = Ita_strike_font

yellow_fill = PatternFill(fill_type="solid",fgColor="FFFF00")
b2.fill = yellow_fill

red2gre = GradientFill(fill_type="linear",stop=("FF0000","00FF00"))
b3.fill = red2gre

thin_side = Side(border_style="thin",color="000000")
double_side = Side(border_style="double",color="FF0000")
b2.border = Border(diagonal=thin_side,diagonalUp=True,diagonalDown=True)

ws.merge_cells("A1:C2")
ws["A1"].value="FishC"
center_align = Alignment(horizontal="center",vertical="center")
ws["A1"].alighment = center_align

highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True,size=20)
highlight.alignment = Alignment(horizontal="center",vertical="center")
wb.add_named_range(highlight)
ws["A1"].style = highlight
ws["B5"].value="LOVE"
ws["B5"].style=highlight

# -------------------------------------------------------------------------------------------

import openpyxl
import datetime
from openpyxl.styles.colors import RED,GREEN,BLUE,YELLOW

wb = openpyxl.Workbook()
ws = wb.active
# 自定义格式
ws["A1"] = 88.8
ws["A1"].number_format = "#,###.00鱼币"

ws["A2"] = datetime.datetime.today()
ws["A2"].number_format = "yyyy-mm-dd"

ws["A3"].number_format = "[RED]+#,###.00;[GREEN]-#,###.00"
ws["A3"].value = 99

ws["A4"].number_format = "[RED];[GREEN];[BLUE];[YELLOW]"
ws["A4"].value = "FishC"

ws["A5"].number_format = "[=1]男;[=0]女"
ws["A5"].value = 1

ws["A6"].number_format = "[<60][RED]不及格;[>=60][GREEN]及格"
ws["A6"].value = 58

wb.save("abc.xlsx")